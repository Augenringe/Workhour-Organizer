import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import re
import shutil
from pathlib import Path

class SimpleExcelProcessor:
    def __init__(self, csv_file_path):
        self.csv_file_path = csv_file_path
        self.month_year = self.extract_month_year_from_filename()
        self.setup_folders()
        self.output_file = os.path.join(self.output_folder, f"Arbeitszeiten_Auswertung_{self.month_year}.xlsx")
        
    def setup_folders(self):
        """Erstellt die notwendigen Ordnerstrukturen"""
        # Hauptordner für die Anwendung (wo die app liegt)
        # Gehe vom aktuellen Skript-Verzeichnis aus
        script_dir = Path(__file__).parent
        self.base_folder = script_dir.parent  # Eine Ebene höher als app/
        
        # Ordner für CSV-Eingabedateien
        self.csv_input_folder = os.path.join(self.base_folder, "CSV_Eingabe")
        
        # Ordner für Excel-Ausgabedateien
        self.output_folder = os.path.join(self.base_folder, "Excel_Ausgabe")
        
        # Ordner für verarbeitete CSV-Dateien (Archiv)
        self.archive_folder = os.path.join(self.base_folder, "CSV_Archiv")
        
        # Erstelle Ordner falls sie nicht existieren
        for folder in [self.csv_input_folder, self.output_folder, self.archive_folder]:
            if not os.path.exists(folder):
                os.makedirs(folder)
                print(f"Ordner erstellt: {folder}")
        
    def extract_month_year_from_filename(self):
        """Extrahiert Monat und Jahr aus dem Dateinamen"""
        filename = os.path.basename(self.csv_file_path)
        
        # Spezielle Behandlung für verschiedene Monatsnamen
        month_patterns = [
            ("Juli", "Juli"), ("August", "August"), ("September", "September"),
            ("Oktober", "Oktober"), ("November", "November"), ("Dezember", "Dezember"),
            ("Januar", "Januar"), ("Februar", "Februar"), ("März", "März"),
            ("April", "April"), ("Mai", "Mai"), ("Juni", "Juni")
        ]
        
        for month_pattern, month_name in month_patterns:
            if month_pattern in filename:
                # Suche nach Jahr im Dateinamen
                year_match = re.search(r'(\d{4})', filename)
                if year_match:
                    return f"{month_name}_{year_match.group(1)}"
                else:
                    return f"{month_name}_{datetime.now().year}"
        
        # Suche nach verschiedenen Datumsformaten im Dateinamen
        patterns = [
            r'(\d{2})\.(\d{4})',  # MM.YYYY
            r'(\d{1,2})\.(\d{4})',  # M.YYYY oder MM.YYYY
            r'(\w+)_(\d{4})',  # MONTH_YYYY
            r'(\d{2})_(\d{4})',  # MM_YYYY
            r'(\d{1,2})_(\d{4})',  # M_YYYY oder MM_YYYY
        ]
        
        for pattern in patterns:
            match = re.search(pattern, filename)
            if match:
                if len(match.group(1)) <= 2:  # Numerischer Monat
                    month_num = int(match.group(1))
                    year = match.group(2)
                    month_names = [
                        "Januar", "Februar", "März", "April", "Mai", "Juni",
                        "Juli", "August", "September", "Oktober", "November", "Dezember"
                    ]
                    month_name = month_names[month_num - 1]
                    return f"{month_name}_{year}"
                else:  # Text-Monat
                    month_name = match.group(1)
                    year = match.group(2)
                    return f"{month_name}_{year}"
        
        # Fallback: Verwende den Dateinamen ohne Erweiterung
        base_name = os.path.splitext(filename)[0]
        return base_name
        
    def archive_csv_file(self):
        """Verschiebt die verarbeitete CSV-Datei ins Archiv"""
        try:
            filename = os.path.basename(self.csv_file_path)
            archive_path = os.path.join(self.archive_folder, filename)
            
            # Falls Datei bereits existiert, füge Zeitstempel hinzu
            if os.path.exists(archive_path):
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                name, ext = os.path.splitext(filename)
                filename = f"{name}_{timestamp}{ext}"
                archive_path = os.path.join(self.archive_folder, filename)
            
            shutil.move(self.csv_file_path, archive_path)
            print(f"CSV-Datei archiviert: {archive_path}")
        except Exception as e:
            print(f"Warnung: Konnte CSV-Datei nicht archivieren: {e}")
        
    def load_csv_data(self):
        """Lädt die CSV-Datei"""
        data = []
        try:
            with open(self.csv_file_path, 'r', encoding='latin-1') as file:
                reader = csv.reader(file, delimiter=';')
                headers = next(reader)  # Erste Zeile sind die Spaltennamen
                
                for row in reader:
                    # Leerzeichen entfernen und prüfen ob genug Spalten vorhanden sind
                    row = [cell.strip() if cell else '' for cell in row]
                    if len(row) >= 13 and row[2] and row[3]:  # Name und Datum müssen vorhanden sein
                        try:
                            data.append({
                                'kreditor': row[0],
                                'personalnummer': row[1],
                                'name': row[2],
                                'datum': row[3],
                                'beginn': row[4],
                                'ende': row[5],
                                'auftrag': row[6],
                                'arbeitsort': row[7],
                                'geraete_einsatz': row[8],
                                'lst_nr': row[9],
                                'std_satz': row[10],
                                'stunden': float(row[11].replace(',', '.')) if row[11] else 0,
                                'gesamt': float(row[12].replace(',', '.')) if row[12] else 0
                            })
                        except ValueError as e:
                            print(f"Fehler beim Konvertieren der Zeile: {row}")
                            continue
            
            print(f"Lade {len(data)} Datensätze...")
            return data
        except Exception as e:
            print(f"Fehler beim Laden der CSV-Datei: {e}")
            return []
    
    def create_excel_workbook(self, data):
        """Erstellt eine Excel-Datei mit verschiedenen Arbeitsblättern"""
        wb = Workbook()
        
        # Entferne das Standard-Arbeitsblatt
        wb.remove(wb.active)
        
        # Erstelle verschiedene Arbeitsblätter
        self.create_raw_data_sheet(wb, data)
        self.create_mitarbeiter_summary_sheet(wb, data)
        self.create_daily_summary_sheet(wb, data)
        self.create_monthly_summary_sheet(wb, data)
        
        # Erstelle individuelle Arbeitsblätter für jeden Mitarbeiter
        self.create_individual_mitarbeiter_sheets(wb, data)
        
        # Speichere die Datei
        wb.save(self.output_file)
        print(f"Excel-Datei erstellt: {self.output_file}")
        
    def create_raw_data_sheet(self, wb, data):
        """Erstellt ein Arbeitsblatt mit den Rohdaten"""
        ws = wb.create_sheet("Rohdaten")
        
        # Überschrift
        ws['A1'] = f"ARBEITSZEITEN ROHDATEN - Agilos QCS GmbH {self.month_year}"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Spaltenüberschriften
        headers = ['Kreditor', 'Pers.-Nr', 'Name', 'Datum', 'Beginn', 'Ende', 
                  'Auftrag', 'Arbeitsort', 'Geräte-Einsatz', 'Lst.-Nr.', 'Std.-Satz', 'Stunden', 'Gesamt']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Daten einfügen
        for row_idx, row_data in enumerate(data, 4):
            ws.cell(row=row_idx, column=1, value=row_data['kreditor'])
            ws.cell(row=row_idx, column=2, value=row_data['personalnummer'])
            ws.cell(row=row_idx, column=3, value=row_data['name'])
            ws.cell(row=row_idx, column=4, value=row_data['datum'])
            ws.cell(row=row_idx, column=5, value=row_data['beginn'])
            ws.cell(row=row_idx, column=6, value=row_data['ende'])
            ws.cell(row=row_idx, column=7, value=row_data['auftrag'])
            ws.cell(row=row_idx, column=8, value=row_data['arbeitsort'])
            ws.cell(row=row_idx, column=9, value=row_data['geraete_einsatz'])
            ws.cell(row=row_idx, column=10, value=row_data['lst_nr'])
            ws.cell(row=row_idx, column=11, value=row_data['std_satz'])
            ws.cell(row=row_idx, column=12, value=row_data['stunden'])
            ws.cell(row=row_idx, column=13, value=row_data['gesamt'])
        
        # Formatierung der Daten
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Spaltenbreiten anpassen
        column_widths = [12, 10, 20, 12, 10, 10, 15, 25, 15, 10, 10, 10, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Filter hinzufügen
        ws.auto_filter.ref = f"A3:M{ws.max_row}"
        
    def create_mitarbeiter_summary_sheet(self, wb, data):
        """Erstellt ein Arbeitsblatt mit Mitarbeiter-Zusammenfassung"""
        ws = wb.create_sheet("Mitarbeiter-Übersicht")
        
        # Überschrift
        ws['A1'] = f"MITARBEITER ÜBERSICHT - {self.month_year}"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Spaltenüberschriften
        headers = ['Personalnummer', 'Name', 'Anzahl Einträge', 'Gesamtstunden', 'Gesamtbetrag (€)', 'Durchschnitt/Stunde']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Mitarbeiter-Daten gruppieren
        mitarbeiter_data = {}
        for row in data:
            key = (row['personalnummer'], row['name'])
            if key not in mitarbeiter_data:
                mitarbeiter_data[key] = {
                    'personalnummer': row['personalnummer'],
                    'name': row['name'],
                    'eintraege': 0,
                    'stunden': 0,
                    'gesamt': 0
                }
            mitarbeiter_data[key]['eintraege'] += 1
            mitarbeiter_data[key]['stunden'] += row['stunden']
            mitarbeiter_data[key]['gesamt'] += row['gesamt']
        
        # Daten einfügen
        row_idx = 4
        for mitarbeiter in mitarbeiter_data.values():
            ws.cell(row=row_idx, column=1, value=mitarbeiter['personalnummer'])
            ws.cell(row=row_idx, column=2, value=mitarbeiter['name'])
            ws.cell(row=row_idx, column=3, value=mitarbeiter['eintraege'])
            ws.cell(row=row_idx, column=4, value=mitarbeiter['stunden'])
            ws.cell(row=row_idx, column=5, value=mitarbeiter['gesamt'])
            
            # Durchschnitt berechnen
            if mitarbeiter['stunden'] > 0:
                durchschnitt = mitarbeiter['gesamt'] / mitarbeiter['stunden']
                ws.cell(row=row_idx, column=6, value=durchschnitt)
            
            row_idx += 1
        
        # Formatierung
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for col, cell in enumerate(row, 1):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if col in [4, 5, 6]:  # Zahlenformat
                    if col == 4:  # Stunden
                        cell.number_format = '0.00'
                    elif col == 5:  # Betrag
                        cell.number_format = '#,##0.00€'
                    elif col == 6:  # Durchschnitt
                        cell.number_format = '#,##0.00€'
        
        # Spaltenbreiten
        column_widths = [15, 25, 15, 15, 18, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Summenzeile
        last_row = ws.max_row + 1
        ws.cell(row=last_row, column=1, value="GESAMT:")
        ws.cell(row=last_row, column=1).font = Font(bold=True)
        ws.cell(row=last_row, column=3, value=f"=SUM(C4:C{last_row-1})")
        ws.cell(row=last_row, column=4, value=f"=SUM(D4:D{last_row-1})")
        ws.cell(row=last_row, column=5, value=f"=SUM(E4:E{last_row-1})")
        
        # Summenzeile formatieren
        for col in range(1, 7):
            cell = ws.cell(row=last_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            if col in [4, 5]:
                cell.number_format = '#,##0.00€'
        
    def create_daily_summary_sheet(self, wb, data):
        """Erstellt ein Arbeitsblatt mit Tages-Zusammenfassung"""
        ws = wb.create_sheet("Tages-Übersicht")
        
        # Überschrift
        ws['A1'] = f"TAGES ÜBERSICHT - {self.month_year}"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Spaltenüberschriften
        headers = ['Datum', 'Anzahl Mitarbeiter', 'Gesamtstunden', 'Gesamtbetrag (€)', 'Durchschnitt/Stunde']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Tages-Daten gruppieren
        daily_data = {}
        for row in data:
            datum = row['datum']
            if datum not in daily_data:
                daily_data[datum] = {
                    'mitarbeiter': set(),
                    'stunden': 0,
                    'gesamt': 0
                }
            daily_data[datum]['mitarbeiter'].add(row['name'])
            daily_data[datum]['stunden'] += row['stunden']
            daily_data[datum]['gesamt'] += row['gesamt']
        
        # Daten einfügen
        row_idx = 4
        for datum, day_data in sorted(daily_data.items()):
            ws.cell(row=row_idx, column=1, value=datum)
            ws.cell(row=row_idx, column=2, value=len(day_data['mitarbeiter']))
            ws.cell(row=row_idx, column=3, value=day_data['stunden'])
            ws.cell(row=row_idx, column=4, value=day_data['gesamt'])
            
            # Durchschnitt berechnen
            if day_data['stunden'] > 0:
                durchschnitt = day_data['gesamt'] / day_data['stunden']
                ws.cell(row=row_idx, column=5, value=durchschnitt)
            
            row_idx += 1
        
        # Formatierung
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for col, cell in enumerate(row, 1):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if col in [3, 4, 5]:  # Zahlenformat
                    if col == 3:  # Stunden
                        cell.number_format = '0.00'
                    elif col == 4:  # Betrag
                        cell.number_format = '#,##0.00€'
                    elif col == 5:  # Durchschnitt
                        cell.number_format = '#,##0.00€'
        
        # Spaltenbreiten
        column_widths = [15, 18, 15, 18, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
    def create_monthly_summary_sheet(self, wb, data):
        """Erstellt ein Arbeitsblatt mit Monats-Zusammenfassung"""
        ws = wb.create_sheet("Monats-Übersicht")
        
        # Überschrift
        ws['A1'] = f"MONATS ÜBERSICHT - {self.month_year}"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Statistiken berechnen
        gesamt_stunden = sum(row['stunden'] for row in data)
        gesamt_betrag = sum(row['gesamt'] for row in data)
        mitarbeiter_anzahl = len(set(row['name'] for row in data))
        eintraege_anzahl = len(data)
        
        stats = [
            ["Gesamtstunden", gesamt_stunden, "Stunden"],
            ["Gesamtbetrag", gesamt_betrag, "€"],
            ["Anzahl Mitarbeiter", mitarbeiter_anzahl, "Stück"],
            ["Anzahl Einträge", eintraege_anzahl, "Stück"],
            ["Durchschnitt Stunden/Mitarbeiter", gesamt_stunden / mitarbeiter_anzahl if mitarbeiter_anzahl > 0 else 0, "Stunden"],
            ["Durchschnitt Betrag/Mitarbeiter", gesamt_betrag / mitarbeiter_anzahl if mitarbeiter_anzahl > 0 else 0, "€"],
            ["Durchschnitt Stunden/Eintrag", gesamt_stunden / eintraege_anzahl if eintraege_anzahl > 0 else 0, "Stunden"],
            ["Durchschnitt Betrag/Eintrag", gesamt_betrag / eintraege_anzahl if eintraege_anzahl > 0 else 0, "€"]
        ]
        
        # Spaltenüberschriften
        headers = ['Kennzahl', 'Wert', 'Einheit']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Daten einfügen
        for idx, (stat_name, stat_value, unit) in enumerate(stats):
            ws.cell(row=idx+4, column=1, value=stat_name)
            ws.cell(row=idx+4, column=2, value=stat_value)
            ws.cell(row=idx+4, column=3, value=unit)
            
            # Zahlenformat
            if "Stunden" in stat_name:
                ws.cell(row=idx+4, column=2).number_format = '0.00'
            elif "Betrag" in stat_name:
                ws.cell(row=idx+4, column=2).number_format = '#,##0.00€'
            elif "Anzahl" in stat_name:
                ws.cell(row=idx+4, column=2).number_format = '0'
        
        # Formatierung
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for col, cell in enumerate(row, 1):
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Spaltenbreiten
        column_widths = [30, 20, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    def create_individual_mitarbeiter_sheets(self, wb, data):
        """Erstellt für jeden Mitarbeiter ein eigenes Arbeitsblatt"""
        # Mitarbeiter gruppieren
        mitarbeiter_data = {}
        for row in data:
            key = (row['personalnummer'], row['name'])
            if key not in mitarbeiter_data:
                mitarbeiter_data[key] = []
            mitarbeiter_data[key].append(row)
        
        # Für jeden Mitarbeiter ein Arbeitsblatt erstellen
        for (personalnummer, name), eintraege in mitarbeiter_data.items():
            # Arbeitsblatt-Name (Excel-Limit: 31 Zeichen)
            sheet_name = f"{name[:20]}" if len(name) > 20 else name
            sheet_name = sheet_name.replace('/', '_').replace('\\', '_').replace('*', '_').replace('?', '_').replace('[', '_').replace(']', '_')
            
            ws = wb.create_sheet(sheet_name)
            
            # Überschrift
            ws['A1'] = f"ARBEITSZEITEN - {name} (Pers.-Nr: {personalnummer}) - {self.month_year}"
            ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Spaltenüberschriften
            headers = ['Datum', 'Beginn', 'Ende', 'Auftrag', 'Arbeitsort', 'Lst.-Nr.', 'Std.-Satz', 'Stunden', 'Gesamt']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Daten einfügen (sortiert nach Datum)
            sorted_eintraege = sorted(eintraege, key=lambda x: x['datum'])
            for row_idx, eintrag in enumerate(sorted_eintraege, 4):
                ws.cell(row=row_idx, column=1, value=eintrag['datum'])
                ws.cell(row=row_idx, column=2, value=eintrag['beginn'])
                ws.cell(row=row_idx, column=3, value=eintrag['ende'])
                ws.cell(row=row_idx, column=4, value=eintrag['auftrag'])
                ws.cell(row=row_idx, column=5, value=eintrag['arbeitsort'])
                ws.cell(row=row_idx, column=6, value=eintrag['lst_nr'])
                ws.cell(row=row_idx, column=7, value=eintrag['std_satz'])
                ws.cell(row=row_idx, column=8, value=eintrag['stunden'])
                ws.cell(row=row_idx, column=9, value=eintrag['gesamt'])
            
            # Formatierung
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
                for col, cell in enumerate(row, 1):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    if col in [7, 8, 9]:  # Zahlenformat
                        if col == 7:  # Std.-Satz
                            cell.number_format = '#,##0.00€'
                        elif col == 8:  # Stunden
                            cell.number_format = '0.00'
                        elif col == 9:  # Gesamt
                            cell.number_format = '#,##0.00€'
            
            # Spaltenbreiten
            column_widths = [12, 10, 10, 15, 25, 10, 10, 10, 12]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # Summenzeile
            last_row = ws.max_row + 1
            ws.cell(row=last_row, column=1, value="GESAMT:")
            ws.cell(row=last_row, column=1).font = Font(bold=True)
            ws.cell(row=last_row, column=8, value=f"=SUM(H4:H{last_row-1})")
            ws.cell(row=last_row, column=9, value=f"=SUM(I4:I{last_row-1})")
            
            # Summenzeile formatieren
            for col in range(1, 10):
                cell = ws.cell(row=last_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                if col in [8, 9]:
                    if col == 8:  # Stunden
                        cell.number_format = '0.00'
                    elif col == 9:  # Gesamt
                        cell.number_format = '#,##0.00€'
            
            # Filter hinzufügen
            ws.auto_filter.ref = f"A3:I{ws.max_row-1}"
    
    def process_data(self):
        """Hauptfunktion zum Verarbeiten der Daten"""
        print(f"Verarbeite Daten für: {self.month_year}")
        data = self.load_csv_data()
        
        if not data:
            print("Keine Daten gefunden!")
            return
        
        print(f"Verarbeite {len(data)} Datensätze...")
        self.create_excel_workbook(data)
        
        # Archiviere die CSV-Datei nach erfolgreicher Verarbeitung
        self.archive_csv_file()
        
        print("Verarbeitung abgeschlossen!")

import sys

def main():
    # CSV-Datei-Pfad aus Kommandozeilenargument oder Standard
    if len(sys.argv) > 1:
        csv_file = sys.argv[1]
    else:
        # Suche nach CSV-Dateien im CSV_Eingabe-Ordner
        # Bestimme den CSV-Eingabe-Ordner relativ zum Skript
        script_dir = Path(__file__).parent
        base_folder = script_dir.parent
        csv_input_folder = os.path.join(base_folder, "CSV_Eingabe")
        if os.path.exists(csv_input_folder):
            csv_files = [f for f in os.listdir(csv_input_folder) if f.endswith('.csv')]
            if csv_files:
                csv_file = os.path.join(csv_input_folder, csv_files[0])
                print(f"Verwende erste gefundene CSV-Datei: {csv_files[0]}")
            else:
                print("Keine CSV-Dateien im CSV_Eingabe-Ordner gefunden!")
                return
        else:
            # Fallback: Suche im aktuellen Verzeichnis
            csv_files = [f for f in os.listdir('.') if f.endswith('.csv')]
            if csv_files:
                csv_file = csv_files[0]
                print(f"Verwende erste gefundene CSV-Datei: {csv_file}")
            else:
                print("Keine CSV-Dateien gefunden!")
                return
    
    if not os.path.exists(csv_file):
        print(f"CSV-Datei nicht gefunden: {csv_file}")
        return
    
    # Verarbeiter erstellen und ausführen
    processor = SimpleExcelProcessor(csv_file)
    processor.process_data()

if __name__ == "__main__":
    main()
